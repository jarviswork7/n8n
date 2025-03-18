import boto3
import openpyxl
from openpyxl import Workbook
import os

# Initialize the EC2 client
client = boto3.client('ec2', region_name='us-east-1')

# Lambda handler
def lambda_handler(event, context):
    
    try:
        # Create EC2 instance
        instance = client.run_instances(
            ImageId='ami-08b5b3a93ed654d19',
            InstanceType='t3.micro',
            MaxCount=1,
            MinCount=1,
            TagSpecifications=[
                {
                    'ResourceType': 'instance',
                    'Tags': [
                        {'Key': 'createdBy', 'Value': 'Daniel'},
                        {'Key': 'Name', 'Value': 'n8n'}
                    ]
                }
            ]
        )

        instance_id = instance['Instances'][0]['InstanceId']

        # Check if Excel file exists
        path = '/tmp/AWS-Resources.xlsx'
        if not os.path.exists(path):
            workbook = Workbook()
            workbook.save(path)

        # Load the Excel workbook
        workbook = openpyxl.load_workbook(path)
        
        # Create a new sheet if test-sheet doesn't exist
        if 'test-sheet' not in workbook.sheetnames:
            workbook.create_sheet('test-sheet')
        sheet = workbook['test-sheet']

        # Add instance ID to Excel
        sheet.append([instance_id])

        # Save the workbook
        workbook.save(path)
        
        return {'Status': 'Success', 'InstanceId': instance_id}
    except Exception as e:
        return {'Status': 'Failed', 'Error': str(e)}
